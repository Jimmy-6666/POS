import io
import json
import re
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash

from pos_app import create_app
from pos_app.database import get_db
from pos_app.services.product_spreadsheet import (
    HEADERS,
    SpreadsheetError,
    apply_import,
    parse_import,
)
from tests.staff_helpers import staff_login


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(list(HEADERS))
    for values in rows:
        sheet.append(values)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def row_values(**values):
    return [values.get(header) for header in HEADERS]


class ProductXlsxTests(unittest.TestCase):
    def setUp(self):
        self.folder = tempfile.TemporaryDirectory()
        self.app = create_app({"TESTING": True, "DATABASE": str(Path(self.folder.name) / "products.xlsx.db")})
        with self.app.app_context():
            db = get_db()
            db.execute("UPDATE staff SET must_change_pin=0 WHERE id=1")
            manager_role = db.execute("SELECT id FROM roles WHERE code='manager'").fetchone()[0]
            db.execute(
                "INSERT INTO staff(display_name,pin_hash,role_id,must_change_pin) VALUES(?,?,?,0)",
                ("Manager", generate_password_hash("2222"), manager_role),
            )
            self.category = db.execute("SELECT name_th FROM categories WHERE is_active=1 ORDER BY id LIMIT 1").fetchone()[0]
            self.unit = db.execute("SELECT name_th FROM units WHERE is_active=1 ORDER BY id LIMIT 1").fetchone()[0]
            category_id = db.execute("SELECT id FROM categories WHERE name_th=?", (self.category,)).fetchone()[0]
            unit_id = db.execute("SELECT id FROM units WHERE name_th=?", (self.unit,)).fetchone()[0]
            db.execute(
                """INSERT INTO products
                   (barcode,sku,name_th,category_id,unit_id,cost_satang,price_satang,stock_quantity)
                   VALUES('00123','SKU-00123','สินค้าเดิม',?,?,100,250,4)""",
                (category_id, unit_id),
            )
            db.commit()
            self.product_uuid = db.execute("SELECT product_uuid FROM products WHERE barcode='00123'").fetchone()[0]
        self.client = self.app.test_client()
        staff_login(self.client)

    def tearDown(self):
        self.folder.cleanup()

    def csrf(self):
        with self.app.app_context():
            return get_db().execute("SELECT csrf_token FROM staff_sessions WHERE staff_id=1").fetchone()[0]

    def upload(self, content, filename="products.xlsx"):
        return self.client.post(
            "/products/xlsx/import",
            data={"csrf_token": self.csrf(), "file": (io.BytesIO(content), filename)},
            content_type="multipart/form-data",
        )

    def confirm(self, preview_response):
        token = re.search(r'name="preview_token" value="([^"]+)"', preview_response.get_data(as_text=True)).group(1)
        return self.client.post(
            "/products/xlsx/import/confirm",
            data={"csrf_token": self.csrf(), "preview_token": token},
        )

    def test_export_is_admin_only_and_is_a_reusable_template(self):
        with self.app.app_context():
            before = tuple(get_db().execute(
                "SELECT barcode,sku,name_th,name_en,cost_satang,price_satang FROM products WHERE product_uuid=?",
                (self.product_uuid,),
            ).fetchone())
        response = self.client.get("/products/xlsx/export")
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response.headers["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(response.data), data_only=False)
        self.assertEqual(workbook["Products"].freeze_panes, "A2")
        self.assertIn("Instructions", workbook.sheetnames)
        self.assertEqual([cell.value for cell in workbook["Products"][1]], list(HEADERS))
        self.assertEqual(workbook["Products"]["A2"].value, self.product_uuid)
        self.assertEqual(workbook["Products"]["C2"].number_format, "@")
        with self.app.app_context():
            after = tuple(get_db().execute(
                "SELECT barcode,sku,name_th,name_en,cost_satang,price_satang FROM products WHERE product_uuid=?",
                (self.product_uuid,),
            ).fetchone())
            self.assertEqual(after, before)
            self.assertEqual(get_db().execute("SELECT COUNT(*) FROM audit_logs WHERE action='product_xlsx_export'").fetchone()[0], 1)

        self.client.post("/logout", data={"csrf_token": self.csrf()})
        manager = self.app.test_client()
        staff_login(manager, 2, "2222")
        self.assertEqual(manager.get("/products/xlsx/export").status_code, 403)
        self.assertEqual(manager.post("/products/xlsx/import", data={}).status_code, 403)

    def test_preview_changes_nothing_then_updates_only_by_product_uuid(self):
        content = workbook_bytes([row_values(product_uuid=self.product_uuid, sku="SKU-CHANGED", barcode="00999", name_th="ชื่อใหม่")])
        preview = self.upload(content)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("ชื่อใหม่", preview.get_data(as_text=True))
        self.assertIn("1.00 บาท", preview.get_data(as_text=True))
        self.assertIn("2.50 บาท", preview.get_data(as_text=True))
        with self.app.app_context():
            row = get_db().execute("SELECT barcode,sku,name_th FROM products WHERE product_uuid=?", (self.product_uuid,)).fetchone()
            self.assertEqual(tuple(row), ("00123", "SKU-00123", "สินค้าเดิม"))
        self.assertEqual(self.confirm(preview).status_code, 302)
        with self.app.app_context():
            row = get_db().execute("SELECT product_uuid,barcode,sku,name_th FROM products WHERE barcode='00999'").fetchone()
            self.assertEqual(tuple(row), (self.product_uuid, "00999", "SKU-CHANGED", "ชื่อใหม่"))
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM audit_logs WHERE action='product_xlsx_import'").fetchone()[0], 1)
            product_audit = db.execute(
                "SELECT entity_id,new_value FROM audit_logs WHERE action='product_xlsx_update'"
            ).fetchone()
            self.assertEqual(product_audit["entity_id"], self.product_uuid)
            self.assertEqual(json.loads(product_audit["new_value"]), {
                "source": "xlsx_import", "changed_fields": ["barcode", "name_th", "sku"],
            })
            self.assertNotIn("ชื่อใหม่", product_audit["new_value"])

    def test_export_round_trip_ignores_existing_trailing_text_whitespace(self):
        with self.app.app_context():
            db = get_db()
            db.execute("UPDATE products SET name_en=? WHERE product_uuid=?", ("Existing English name\n", self.product_uuid))
            db.commit()
        exported = self.client.get("/products/xlsx/export")
        preview = self.upload(exported.data)
        body = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("แก้ไข</strong><p>0", body)
        self.assertIn("ไม่มีรายละเอียดแถวด้านล่าง", body)

    def test_missing_uuid_creates_product_with_audited_opening_stock(self):
        content = workbook_bytes([row_values(
            barcode="00007", name_th="สินค้าใหม่", category=self.category, unit=self.unit,
            cost_price_baht=12.5, selling_price_baht=20, stock_quantity=3, active="yes",
        )])
        preview = self.upload(content)
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(self.confirm(preview).status_code, 302)
        with self.app.app_context():
            db = get_db()
            product = db.execute("SELECT product_uuid,stock_quantity FROM products WHERE barcode='00007'").fetchone()
            self.assertRegex(product["product_uuid"], r"^[0-9a-f-]{36}$")
            self.assertEqual(product["stock_quantity"], 3)
            movement = db.execute("SELECT movement_type,changed_quantity FROM stock_movements WHERE product_id=(SELECT id FROM products WHERE barcode='00007')").fetchone()
            self.assertEqual(tuple(movement), ("opening_balance", 3))
            product_audit = db.execute(
                "SELECT entity_id,new_value FROM audit_logs WHERE action='product_xlsx_create'"
            ).fetchone()
            self.assertEqual(product_audit["entity_id"], product["product_uuid"])
            self.assertNotIn("สินค้าใหม่", product_audit["new_value"])

    def test_invalid_uuid_and_formula_are_rejected(self):
        content = workbook_bytes([
            row_values(product_uuid="not-a-uuid", barcode="00124", name_th="ผิด", category=self.category, unit=self.unit, cost_price_baht=1, selling_price_baht=2),
            row_values(product_uuid=self.product_uuid, barcode="=DANGEROUS"),
        ])
        preview = self.upload(content)
        body = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("Product UUID", body)
        self.assertIn("Formula-like", body)

    def test_existing_negative_stock_is_ignored_while_catalog_fields_update(self):
        with self.app.app_context():
            db = get_db()
            db.execute(
                "UPDATE products SET stock_quantity=-3 WHERE product_uuid=?",
                (self.product_uuid,),
            )
            db.commit()

        content = workbook_bytes([
            row_values(
                product_uuid=self.product_uuid,
                name_th="แก้ชื่อแม้สต็อกติดลบ",
                stock_quantity=-3,
            ),
        ])
        preview = self.upload(content)
        body = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertNotIn("Stock quantity", body)
        self.assertEqual(self.confirm(preview).status_code, 302)
        with self.app.app_context():
            db = get_db()
            product = db.execute(
                "SELECT name_th,stock_quantity FROM products WHERE product_uuid=?",
                (self.product_uuid,),
            ).fetchone()
            self.assertEqual(tuple(product), ("แก้ชื่อแม้สต็อกติดลบ", -3))
            audit = db.execute(
                "SELECT new_value FROM audit_logs WHERE action='product_xlsx_update'"
            ).fetchone()
            self.assertEqual(
                json.loads(audit["new_value"])["changed_fields"],
                ["name_th"],
            )

    def test_repeated_product_uuid_is_rejected_in_preview(self):
        content = workbook_bytes([
            row_values(product_uuid=self.product_uuid, name_th="ชื่อหนึ่ง"),
            row_values(product_uuid=self.product_uuid, name_th="ชื่อสอง"),
        ])
        preview = self.upload(content)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("repeated", preview.get_data(as_text=True))

    def test_transaction_rolls_back_when_data_changes_after_preview(self):
        content = workbook_bytes([
            row_values(barcode="NEW-ATOMIC", name_th="รายการแรก", category=self.category, unit=self.unit, cost_price_baht=1, selling_price_baht=2),
            row_values(barcode="LATE-COLLISION", name_th="รายการสอง", category=self.category, unit=self.unit, cost_price_baht=1, selling_price_baht=2),
        ])
        with self.app.app_context():
            db = get_db()
            preview = parse_import(db, content)
            self.assertFalse(preview["errors"])
            category_id = db.execute("SELECT id FROM categories WHERE name_th=?", (self.category,)).fetchone()[0]
            unit_id = db.execute("SELECT id FROM units WHERE name_th=?", (self.unit,)).fetchone()[0]
            db.execute("INSERT INTO products(barcode,name_th,category_id,unit_id,cost_satang,price_satang) VALUES('LATE-COLLISION','คู่แข่ง',?,?,1,2)", (category_id, unit_id))
            db.commit()
            with self.assertRaises(Exception):
                apply_import(db, 1, "products.xlsx", preview, "127.0.0.1")
            self.assertIsNone(db.execute("SELECT 1 FROM products WHERE barcode='NEW-ATOMIC'").fetchone())

    def test_revalidation_rejects_deleted_update_target_and_rolls_back_batch(self):
        content = workbook_bytes([
            row_values(product_uuid=self.product_uuid, name_th="จะไม่ถูกบันทึก"),
            row_values(barcode="NEW-REVALIDATE", name_th="ต้อง rollback", category=self.category, unit=self.unit, cost_price_baht=1, selling_price_baht=2),
        ])
        with self.app.app_context():
            db = get_db()
            preview = parse_import(db, content)
            self.assertFalse(preview["errors"])
            db.execute("DELETE FROM products WHERE product_uuid=?", (self.product_uuid,))
            db.commit()
            with self.assertRaisesRegex(SpreadsheetError, "no longer exists"):
                apply_import(db, 1, "products.xlsx", preview, "127.0.0.1")
            self.assertIsNone(db.execute("SELECT 1 FROM products WHERE barcode='NEW-REVALIDATE'").fetchone())
            self.assertEqual(
                db.execute("SELECT COUNT(*) FROM audit_logs WHERE action IN ('product_xlsx_create','product_xlsx_update')").fetchone()[0],
                0,
            )


if __name__ == "__main__":
    unittest.main()
