"""Admin-only XLSX product interchange.

The spreadsheet is deliberately an interchange format, not a second product
API.  ``product_uuid`` is the only update key; barcode and SKU stay editable
business values and are never used to find an existing product.
"""

from __future__ import annotations

import json
import secrets
import threading
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from werkzeug.utils import secure_filename

from ..product_identity import ProductIdentityError, canonical_product_uuid, new_product_uuid
from .money import baht_to_satang, satang_to_baht


MAX_IMPORT_BYTES = 2 * 1024 * 1024
PREVIEW_TTL = timedelta(minutes=15)
CLEAR_TOKEN = "__CLEAR__"
SHEET_NAME = "Products"
INSTRUCTIONS_SHEET = "Instructions"
REFERENCE_SHEET = "_Reference values"

HEADERS = (
    "product_uuid", "sku", "barcode", "name_th", "name_en", "category", "unit",
    "cost_price_baht", "selling_price_baht", "stock_quantity", "minimum_stock",
    "active", "favorite", "allow_decimal_quantity", "online_available",
    "online_sort_order", "online_max_quantity",
)

HEADER_LABELS = {
    "product_uuid": "Product UUID", "sku": "SKU", "barcode": "Barcode", "name_th": "Thai name",
    "name_en": "English name", "category": "Category", "unit": "Unit",
    "cost_price_baht": "Cost price (THB)", "selling_price_baht": "Selling price (THB)",
    "stock_quantity": "Stock quantity", "minimum_stock": "Minimum stock", "active": "Active",
    "favorite": "Favorite", "allow_decimal_quantity": "Allow decimal quantity",
    "online_available": "Online available", "online_sort_order": "Online sort order",
    "online_max_quantity": "Online max quantity",
}

TEXT_FIELDS = {"sku", "barcode", "name_th", "name_en", "category", "unit"}
BOOLEAN_FIELDS = {"active", "favorite", "allow_decimal_quantity", "online_available"}
NULLABLE_FIELDS = {"sku", "name_en", "online_max_quantity"}
FORMULA_PREFIXES = ("=", "+", "-", "@")


class SpreadsheetError(ValueError):
    """A user-safe spreadsheet validation error."""


@dataclass(frozen=True)
class StoredPreview:
    staff_id: int
    expires_at: datetime
    filename: str
    preview: dict[str, Any]


_previews: dict[str, StoredPreview] = {}
_preview_lock = threading.Lock()


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _clean_previews() -> None:
    now = _now()
    for token, value in list(_previews.items()):
        if value.expires_at <= now:
            _previews.pop(token, None)


def remember_preview(staff_id: int, filename: str, preview: dict[str, Any]) -> str:
    """Keep a parsed, non-persistent preview briefly until explicit confirmation."""

    token = secrets.token_urlsafe(24)
    with _preview_lock:
        _clean_previews()
        _previews[token] = StoredPreview(staff_id, _now() + PREVIEW_TTL, filename, preview)
    return token


def take_preview(staff_id: int, token: str) -> StoredPreview:
    with _preview_lock:
        _clean_previews()
        preview = _previews.get(token)
        if not preview or preview.staff_id != staff_id or preview.expires_at <= _now():
            raise SpreadsheetError("ตัวอย่างการนำเข้าหมดอายุแล้ว กรุณาอัปโหลดไฟล์อีกครั้ง")
        # A preview is intentionally one-time: a double submit cannot replay an import.
        return _previews.pop(token)


def safe_filename(filename: str | None) -> str:
    return secure_filename(filename or "") or "products.xlsx"


def _safe_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip() or None


def _has_formula(value: Any) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES)


def _export_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return f"'{text}" if text.startswith(FORMULA_PREFIXES) else text


def _required(value: Any, column: str) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise SpreadsheetError(f"{HEADER_LABELS[column]} is required")
    return value


def _decimal(value: Any, column: str, *, whole: bool = False) -> float:
    try:
        decimal = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise SpreadsheetError(f"{HEADER_LABELS[column]} must be a number") from exc
    if not decimal.is_finite() or decimal < 0:
        raise SpreadsheetError(f"{HEADER_LABELS[column]} cannot be negative")
    if whole and decimal != decimal.to_integral_value():
        raise SpreadsheetError(f"{HEADER_LABELS[column]} must be a whole number")
    return float(decimal)


def _boolean(value: Any, column: str) -> int:
    if isinstance(value, bool):
        return int(value)
    text = _safe_text(value)
    if text is None:
        raise SpreadsheetError(f"{HEADER_LABELS[column]} must be yes or no")
    normalized = text.casefold()
    if normalized in {"yes", "true", "1", "y", "ใช่"}:
        return 1
    if normalized in {"no", "false", "0", "n", "ไม่"}:
        return 0
    raise SpreadsheetError(f"{HEADER_LABELS[column]} must be yes or no")


def _reference_data(db):
    categories = db.execute("SELECT id,name_th FROM categories WHERE is_active=1 ORDER BY sort_order,name_th").fetchall()
    units = db.execute("SELECT id,name_th FROM units WHERE is_active=1 ORDER BY sort_order,name_th").fetchall()
    return categories, units


def _reference_map(rows):
    return {str(row["name_th"]).strip().casefold(): (row["id"], row["name_th"]) for row in rows}


def _parse_reference(value: Any, column: str, lookup: dict[str, tuple[int, str]]) -> tuple[int, str]:
    text = _safe_text(_required(value, column))
    if _has_formula(text):
        raise SpreadsheetError(f"{HEADER_LABELS[column]} cannot begin with a spreadsheet formula character")
    match = lookup.get(text.casefold())
    if not match:
        raise SpreadsheetError(f"{HEADER_LABELS[column]} is not an active configured value")
    return match


def _same_import_value(column: str, existing: Any, incoming: Any) -> bool:
    """Avoid a no-op update when Excel round-trips harmless text whitespace."""

    if column in TEXT_FIELDS:
        return _safe_text(existing) == _safe_text(incoming)
    return existing == incoming


def _preview_product_details(data: dict[str, Any], current=None) -> dict[str, str]:
    """Return display-only final product information without retaining sheet rows."""

    values = dict(current) if current is not None else {}
    values.update(data)
    return {
        "name_th": values["name_th"],
        "cost_price_baht": satang_to_baht(values["cost_satang"]),
        "selling_price_baht": satang_to_baht(values["price_satang"]),
    }


def _parse_row(db, row_number: int, values: dict[str, Any], categories, units):
    """Return an import action or detailed per-column errors for one row."""

    errors: list[dict[str, Any]] = []

    def field(column: str, parser, *, required: bool = False):
        raw = values.get(column)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            if required:
                errors.append({"row": row_number, "column": column, "reason": f"{HEADER_LABELS[column]} is required"})
            return None, False
        if isinstance(raw, str) and raw.strip() == CLEAR_TOKEN:
            if column in NULLABLE_FIELDS:
                return None, True
            errors.append({"row": row_number, "column": column, "reason": "This field cannot be cleared"})
            return None, False
        try:
            if _has_formula(raw):
                raise SpreadsheetError("Formula-like values are not allowed")
            return parser(raw, column), True
        except SpreadsheetError as exc:
            errors.append({"row": row_number, "column": column, "reason": str(exc)})
            return None, False

    raw_uuid = values.get("product_uuid")
    action_type = "create"
    current = None
    product_uuid = None
    if raw_uuid is not None and (not isinstance(raw_uuid, str) or raw_uuid.strip()):
        try:
            if _has_formula(raw_uuid):
                raise ProductIdentityError("formula")
            product_uuid = canonical_product_uuid(_safe_text(raw_uuid))
            current = db.execute("SELECT * FROM products WHERE product_uuid=?", (product_uuid,)).fetchone()
            if not current:
                raise SpreadsheetError("Product UUID does not exist; leave it blank to create a new product")
            action_type = "update"
        except ProductIdentityError:
            errors.append({"row": row_number, "column": "product_uuid", "reason": "Product UUID must be a valid existing UUID"})
        except SpreadsheetError as exc:
            errors.append({"row": row_number, "column": "product_uuid", "reason": str(exc)})

    category_lookup = _reference_map(categories)
    unit_lookup = _reference_map(units)
    parsed: dict[str, Any] = {}

    for column in ("sku", "barcode", "name_th", "name_en"):
        required = action_type == "create" and column in {"barcode", "name_th"}
        parsed_value, supplied = field(
            column,
            lambda value, key: _safe_text(value) if not _has_formula(_safe_text(value)) else (_ for _ in ()).throw(SpreadsheetError("Formula-like values are not allowed")),
            required=required,
        )
        if supplied:
            parsed[column] = parsed_value

    for column, lookup in (("category", category_lookup), ("unit", unit_lookup)):
        required = action_type == "create"
        raw = values.get(column)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            if required:
                errors.append({"row": row_number, "column": column, "reason": f"{HEADER_LABELS[column]} is required"})
        elif isinstance(raw, str) and raw.strip() == CLEAR_TOKEN:
            errors.append({"row": row_number, "column": column, "reason": "This field cannot be cleared"})
        else:
            try:
                identifier, _ = _parse_reference(raw, column, lookup)
                parsed[f"{column}_id"] = identifier
            except SpreadsheetError as exc:
                errors.append({"row": row_number, "column": column, "reason": str(exc)})

    for column in ("cost_price_baht", "selling_price_baht"):
        required = action_type == "create"
        raw = values.get(column)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            if required:
                errors.append({"row": row_number, "column": column, "reason": f"{HEADER_LABELS[column]} is required"})
        else:
            try:
                if _has_formula(raw):
                    raise SpreadsheetError("Formula-like values are not allowed")
                parsed[{"cost_price_baht": "cost_satang", "selling_price_baht": "price_satang"}[column]] = baht_to_satang(raw)
            except (SpreadsheetError, ValueError) as exc:
                errors.append({"row": row_number, "column": column, "reason": str(exc)})

    for column in ("minimum_stock", "online_max_quantity"):
        raw = values.get(column)
        if raw is not None and (not isinstance(raw, str) or raw.strip()):
            if isinstance(raw, str) and raw.strip() == CLEAR_TOKEN and column in NULLABLE_FIELDS:
                parsed[column] = None
            else:
                try:
                    if _has_formula(raw):
                        raise SpreadsheetError("Formula-like values are not allowed")
                    parsed[column] = _decimal(raw, column)
                except SpreadsheetError as exc:
                    errors.append({"row": row_number, "column": column, "reason": str(exc)})

    raw_sort = values.get("online_sort_order")
    if raw_sort is not None and (not isinstance(raw_sort, str) or raw_sort.strip()):
        try:
            if _has_formula(raw_sort):
                raise SpreadsheetError("Formula-like values are not allowed")
            parsed["online_sort_order"] = int(_decimal(raw_sort, "online_sort_order", whole=True))
        except SpreadsheetError as exc:
            errors.append({"row": row_number, "column": "online_sort_order", "reason": str(exc)})

    for column in BOOLEAN_FIELDS:
        raw = values.get(column)
        if raw is not None and (not isinstance(raw, str) or raw.strip()):
            try:
                if _has_formula(raw):
                    raise SpreadsheetError("Formula-like values are not allowed")
                database_column = {
                    "active": "is_active",
                    "favorite": "is_favorite",
                    "allow_decimal_quantity": "allow_decimal_quantity",
                    "online_available": "is_online_available",
                }[column]
                parsed[database_column] = _boolean(raw, column)
            except SpreadsheetError as exc:
                errors.append({"row": row_number, "column": column, "reason": str(exc)})

    raw_stock = values.get("stock_quantity")
    opening_stock = None
    if raw_stock is not None and (not isinstance(raw_stock, str) or raw_stock.strip()):
        try:
            if _has_formula(raw_stock):
                raise SpreadsheetError("Formula-like values are not allowed")
            opening_stock = _decimal(raw_stock, "stock_quantity")
            if action_type == "update" and current and Decimal(str(opening_stock)) != Decimal(str(current["stock_quantity"])):
                raise SpreadsheetError("Stock quantity is read-only for existing products; use stock adjustment instead")
        except SpreadsheetError as exc:
            errors.append({"row": row_number, "column": "stock_quantity", "reason": str(exc)})

    if errors:
        return None, errors

    if action_type == "create":
        if not parsed.get("allow_decimal_quantity", 0) and opening_stock is not None and not float(opening_stock).is_integer():
            return None, [{"row": row_number, "column": "stock_quantity", "reason": "Decimal stock needs Allow decimal quantity = yes"}]
        data = {
            "product_uuid": new_product_uuid(), "sku": None, "name_en": None,
            "minimum_stock": 0.0, "is_active": 1, "is_favorite": 0,
            "allow_decimal_quantity": 0, "is_online_available": 0,
            "online_sort_order": 0, "online_max_quantity": None,
            "stock_quantity": opening_stock or 0.0,
            **parsed,
        }
        return {
            "kind": "create", "row": row_number, "product_uuid": data["product_uuid"],
            "data": data, "product": _preview_product_details(data),
        }, []

    changes = {}
    for key, value in parsed.items():
        if not _same_import_value(key, current[key], value):
            changes[key] = value
    return {
        "kind": "update" if changes else "unchanged", "row": row_number,
        "product_uuid": product_uuid, "data": changes,
        "product": _preview_product_details(changes, current),
    }, []


def _validate_unique_business_values(db, actions: list[dict[str, Any]], errors: list[dict[str, Any]]) -> None:
    """Catch ordinary SKU/barcode collisions before confirmation (DB enforces again)."""

    planned: dict[str, dict[str, int]] = {"barcode": {}, "sku": {}}
    for action in actions:
        if action["kind"] == "unchanged":
            continue
        data = action["data"]
        for column in ("barcode", "sku"):
            value = data.get(column)
            if not value:
                continue
            owner = db.execute(f"SELECT product_uuid FROM products WHERE {column}=?", (value,)).fetchone()
            if owner and owner["product_uuid"] != action["product_uuid"]:
                errors.append({"row": action["row"], "column": column, "reason": f"{HEADER_LABELS[column]} already belongs to another product"})
            previous_row = planned[column].get(value)
            if previous_row is not None:
                errors.append({"row": action["row"], "column": column, "reason": f"{HEADER_LABELS[column]} is repeated on row {previous_row}"})
            planned[column][value] = action["row"]


def parse_import(db, content: bytes) -> dict[str, Any]:
    if len(content) > MAX_IMPORT_BYTES:
        raise SpreadsheetError("The XLSX file is larger than 2 MB")
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if "[Content_Types].xml" not in archive.namelist() or len(members) > 1_000:
                raise SpreadsheetError("The uploaded file is not a valid XLSX workbook")
            if sum(member.file_size for member in members) > 20 * 1024 * 1024:
                raise SpreadsheetError("The XLSX file expands beyond the safe import limit")
    except BadZipFile as exc:
        raise SpreadsheetError("The uploaded file is not a valid XLSX workbook") from exc
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise SpreadsheetError("The uploaded file is not a valid XLSX workbook") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise SpreadsheetError("The workbook must contain a Products worksheet")
    sheet = workbook[SHEET_NAME]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
    if headers != list(HEADERS):
        raise SpreadsheetError("The Products header row does not match the current export template")

    categories, units = _reference_data(db)
    actions: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_uuids: dict[str, int] = {}
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {header: cell.value for header, cell in zip(HEADERS, cells)}
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in values.values()):
            continue
        action, row_errors = _parse_row(db, row_number, values, categories, units)
        if row_errors:
            errors.extend(row_errors)
            continue
        if action["kind"] != "create":
            repeated = seen_uuids.get(action["product_uuid"])
            if repeated is not None:
                errors.append({"row": row_number, "column": "product_uuid", "reason": f"Product UUID is repeated on row {repeated}"})
                continue
            seen_uuids[action["product_uuid"]] = row_number
        actions.append(action)
    _validate_unique_business_values(db, actions, errors)
    summaries = {kind: [action for action in actions if action["kind"] == kind] for kind in ("create", "update", "unchanged")}
    return {
        "actions": actions,
        "errors": errors,
        "summary": {"created": len(summaries["create"]), "updated": len(summaries["update"]), "unchanged": len(summaries["unchanged"]), "rejected": len(errors)},
        "rows": summaries,
    }


def _revalidate_actions_under_lock(db, actions: list[dict[str, Any]]) -> None:
    """Confirm that preview targets and references still exist after locking writes."""

    errors: list[dict[str, Any]] = []
    for action in actions:
        if action["kind"] == "create":
            if db.execute("SELECT 1 FROM products WHERE product_uuid=?", (action["product_uuid"],)).fetchone():
                errors.append({"row": action["row"], "column": "product_uuid", "reason": "Generated Product UUID is no longer available"})
        elif action["kind"] == "update":
            if not db.execute("SELECT 1 FROM products WHERE product_uuid=?", (action["product_uuid"],)).fetchone():
                errors.append({"row": action["row"], "column": "product_uuid", "reason": "Product UUID no longer exists"})
        for table, column in (("categories", "category_id"), ("units", "unit_id")):
            identifier = action["data"].get(column)
            if identifier is not None and not db.execute(
                f"SELECT 1 FROM {table} WHERE id=? AND is_active=1", (identifier,)
            ).fetchone():
                errors.append({"row": action["row"], "column": column, "reason": "Referenced value is no longer active"})
    _validate_unique_business_values(db, actions, errors)
    if errors:
        first = errors[0]
        raise SpreadsheetError(
            f"Import data changed after preview (row {first['row']}, {first['column']}: {first['reason']}). Upload the file again."
        )


def _record_product_import_audit(db, staff_id: int, action: dict[str, Any], ip_address: str | None, now: str) -> None:
    """Record one non-sensitive audit event for each product that actually changed."""

    changed_fields = sorted(field for field in action["data"] if field != "product_uuid")
    db.execute(
        """INSERT INTO audit_logs (staff_id,action,entity_type,entity_id,new_value,ip_address,created_at)
           VALUES (?, ?, 'product', ?, ?, ?, ?)""",
        (
            staff_id,
            f"product_xlsx_{action['kind']}",
            action["product_uuid"],
            json.dumps({"source": "xlsx_import", "changed_fields": changed_fields}, ensure_ascii=False),
            ip_address,
            now,
        ),
    )


def apply_import(db, staff_id: int, filename: str, preview: dict[str, Any], ip_address: str | None) -> dict[str, int]:
    if preview["errors"]:
        raise SpreadsheetError("Resolve all invalid rows before confirming this import")
    actions = preview["actions"]
    now = _now().isoformat(timespec="seconds")
    try:
        db.execute("BEGIN IMMEDIATE")
        _revalidate_actions_under_lock(db, actions)
        for action in actions:
            data = action["data"]
            if action["kind"] == "create":
                cursor = db.execute(
                    """INSERT INTO products
                       (product_uuid,barcode,sku,name_th,name_en,category_id,unit_id,cost_satang,price_satang,
                        stock_quantity,minimum_stock,is_active,is_favorite,allow_decimal_quantity,is_online_available,
                        online_sort_order,online_max_quantity)
                       VALUES (:product_uuid,:barcode,:sku,:name_th,:name_en,:category_id,:unit_id,:cost_satang,:price_satang,
                               :stock_quantity,:minimum_stock,:is_active,:is_favorite,:allow_decimal_quantity,:is_online_available,
                               :online_sort_order,:online_max_quantity)""",
                    data,
                )
                if data["stock_quantity"]:
                    db.execute(
                        """INSERT INTO stock_movements
                           (product_id,movement_type,previous_quantity,changed_quantity,new_quantity,unit_cost_satang,
                            reference_type,reference_number,staff_id,note,created_at)
                           VALUES (?, 'opening_balance', 0, ?, ?, ?, 'product', ?, ?, 'XLSX opening balance', ?)""",
                        (cursor.lastrowid, data["stock_quantity"], data["stock_quantity"], data["cost_satang"],
                         data["product_uuid"], staff_id, now),
                    )
                _record_product_import_audit(db, staff_id, action, ip_address, now)
            elif action["kind"] == "update":
                set_sql = ", ".join(f"{field}=?" for field in data)
                cursor = db.execute(
                    f"UPDATE products SET {set_sql},updated_at=CURRENT_TIMESTAMP WHERE product_uuid=?",
                    (*data.values(), action["product_uuid"]),
                )
                if cursor.rowcount != 1:
                    raise SpreadsheetError("Product UUID no longer exists; no changes were saved")
                _record_product_import_audit(db, staff_id, action, ip_address, now)
        summary = preview["summary"]
        db.execute(
            """INSERT INTO audit_logs (staff_id,action,entity_type,entity_id,new_value,ip_address,created_at)
               VALUES (?, 'product_xlsx_import', 'product_import', ?, ?, ?, ?)""",
            (staff_id, filename, json.dumps({key: summary[key] for key in ("created", "updated", "rejected")}), ip_address, now),
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return preview["summary"]


def build_export_workbook(db) -> bytes:
    """Return an editable, validated XLSX workbook using current catalog values."""

    categories, units = _reference_data(db)
    products = db.execute(
        """SELECT p.*,c.name_th AS category_name,u.name_th AS unit_name
           FROM products p JOIN categories c ON c.id=p.category_id JOIN units u ON u.id=p.unit_id
           ORDER BY p.name_th,p.id"""
    ).fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.append(list(HEADERS))
    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for product in products:
        sheet.append([
            product["product_uuid"], _export_text(product["sku"]), _export_text(product["barcode"]),
            _export_text(product["name_th"]), _export_text(product["name_en"]), product["category_name"], product["unit_name"],
            float(Decimal(product["cost_satang"]) / 100), float(Decimal(product["price_satang"]) / 100),
            product["stock_quantity"], product["minimum_stock"], "yes" if product["is_active"] else "no",
            "yes" if product["is_favorite"] else "no", "yes" if product["allow_decimal_quantity"] else "no",
            "yes" if product["is_online_available"] else "no", product["online_sort_order"], product["online_max_quantity"],
        ])
    widths = [38, 20, 20, 32, 28, 24, 16, 18, 20, 16, 16, 12, 12, 22, 18, 18, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column in (1, 2, 3):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = "@"
    for column in (8, 9):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = '#,##0.00'
    for column in (10, 11, 17):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = '#,##0.###'
    sheet.auto_filter.ref = f"A1:Q{max(sheet.max_row, 2)}"

    instructions = workbook.create_sheet(INSTRUCTIONS_SHEET)
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 100
    instructions.append(["Product XLSX import instructions", ""])
    instructions["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    instructions["A1"].fill = PatternFill("solid", fgColor="0F766E")
    instructions.merge_cells("A1:B1")
    notes = [
        ("Identity", "product_uuid is the only update key. Keep it for an update. Leave it blank to create a new product. An unknown or malformed UUID is rejected."),
        ("Required for new", "barcode, name_th, category, unit, cost_price_baht, and selling_price_baht."),
        ("Optional", "sku, name_en, minimum_stock, status and online fields. Blank optional cells do not overwrite an existing product."),
        ("Clear value", f"Type {CLEAR_TOKEN} only for SKU, English name, or online max quantity when intentionally clearing it."),
        ("SKU and barcode", "They are editable fields, remain unique, and never identify a product during import. Keep them as text to preserve leading zeroes."),
        ("Stock safety", "For an existing product stock_quantity is read-only; use stock adjustment/receiving. For a new product it becomes an audited opening balance."),
        ("Accepted status", "yes/no, true/false, 1/0, or Thai ใช่/ไม่. Category and unit must be active configured values."),
        ("Safety", "Do not enter formulas or text starting with =, +, - or @. The preview changes nothing; confirmation commits all valid rows atomically."),
    ]
    for label, text in notes:
        instructions.append([label, text])
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True, color="0F172A")
        row[0].fill = PatternFill("solid", fgColor="CCFBF1")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.freeze_panes = "A2"

    references = workbook.create_sheet(REFERENCE_SHEET)
    references.sheet_state = "hidden"
    references.append(["Categories", "Units", "Boolean values"])
    max_reference_rows = max(len(categories), len(units), 2)
    for index in range(max_reference_rows):
        references.append([
            categories[index]["name_th"] if index < len(categories) else None,
            units[index]["name_th"] if index < len(units) else None,
            "yes" if index == 0 else "no" if index == 1 else None,
        ])
    validation_rows = max(500, sheet.max_row + 100)
    for column, formula in (("F", f"'{REFERENCE_SHEET}'!$A$2:$A${len(categories) + 1}"), ("G", f"'{REFERENCE_SHEET}'!$B$2:$B${len(units) + 1}"), ("L", f"'{REFERENCE_SHEET}'!$C$2:$C$3"), ("M", f"'{REFERENCE_SHEET}'!$C$2:$C$3"), ("N", f"'{REFERENCE_SHEET}'!$C$2:$C$3"), ("O", f"'{REFERENCE_SHEET}'!$C$2:$C$3")):
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}{validation_rows}")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_audit_payload(db, staff_id: int, filename: str, ip_address: str | None) -> None:
    count = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    db.execute(
        """INSERT INTO audit_logs (staff_id,action,entity_type,entity_id,new_value,ip_address,created_at)
           VALUES (?, 'product_xlsx_export', 'product_export', ?, ?, ?, ?)""",
        (staff_id, filename, str({"products": count}), ip_address, _now().isoformat(timespec="seconds")),
    )
    db.commit()
